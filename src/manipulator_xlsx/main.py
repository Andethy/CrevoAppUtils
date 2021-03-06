# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
from sys import argv
import os
import _thread
from tkinter import filedialog
from tkinter import *
from typing import TextIO, Any

import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class App:
    """Application class to manage the window and handle pertinent events."""

    # Instantiate the window and instance as class variables
    window = None
    current_inst = None
    dpi = 0.14

    def __init__(self, *args):
        self.args = args

        self.window = Tk()
        App.window = self.window
        # Configure window dimensions & basic formatting
        self.m_screen = ScreenManager()
        self.dim_x = 500
        self.dim_y = 300

        self.m_file = FileIO()

        self.m_csv = FileIO()
        self.m_csv.set_file('data/output.csv')

        self.m_excel = ExcelIO()

        self.m_screen.add_screen("Config",
                                 [LabelCustom(self.window, "Load Excel File"), 0.5, 0.25],
                                 [FileButtonCustom(self.window, self.validate_form, data_type='xlsx'), 0.5, 0.5])
        # [LabelCustom(self.window, "Load Excel File"), 0.5, 0.25],
        # [FileButtonCustom(self.window, self.validate_form, data_type='xlsx'), 0.5, 0.5])
        self.m_screen.add_screen("Home",
                                 [LabelCustom(self.window, "Upload Text File"), 0.5, 0.25],
                                 [FileButtonCustom(self.window, self.validate_input, self.invalidate_input), 0.5, 0.5],
                                 [LabelCustom(self.window, d_width=50, d_height=3), 0.5, 0.75])

        self.run()

    def callback(self):
        self.window.quit()
        _thread.interrupt_main()
        try:
            os._exit(0)
        except InterruptedError:
            print("NORMAL QUIT FAILED.")
            sys.exit()
        finally:
            sys.exit()

    def run(self):
        self.window.title('Scouting Data Transfer')
        App.dpi = App.window.winfo_fpixels('1i')
        print("DPI:", App.dpi)
        self.dim_x = int(self.dim_x * App.dpi) // 96
        self.dim_y = int(self.dim_y * App.dpi) // 96
        self.window.geometry(str(self.dim_x) + 'x' + str(self.dim_y))
        self.window.config(background="black")
        self.window.minsize(width=self.dim_x, height=self.dim_y)
        self.window.maxsize(width=self.dim_x, height=self.dim_y)
        print(os.path.abspath(os.path.curdir))
        photo = PhotoImage(file=resource_path('resources/logo.png'))
        self.window.iconphoto(False, photo)

        App.current_inst = self

        self.m_screen.get_screen("Config").set_screen()

        App.window.mainloop()

    def validate_form(self, file_path):
        self.m_excel.set_workbook(file_path)
        self.m_screen.get_screen("Home").set_screen()

    def validate_input(self, file_path):
        self.m_screen.get_screen("Home").widgets[2].set_value("...")
        self.m_file.set_file(str(file_path))
        self.m_file.set_data_txt()
        self.m_csv.manipulate_file(self.m_file.get_text())
        self.m_excel.modify_spreadsheet(self.m_file.get_data())
        self.m_screen.get_screen("Home").widgets[2].set_value(str("Successfully converted\n"
                                                                  + os.path.basename(self.m_file.file_path)))

    def invalidate_input(self):
        self.m_screen.get_screen("Home").widgets[2].set_value("Error")


class FileIO:
    file: TextIO
    file_path: str
    data: list[tuple]

    def __init__(self):
        self.file = ...
        self.file_path = ''
        self.data = []

    def set_file(self, file_path):
        self.file_path = str(file_path)
        self.file = open(str(self.file_path))

    def set_data_txt(self):
        self.data = []
        lines = self.file.read().splitlines()
        for line in lines:
            print(line, lines)
            if line == '':
                continue
            data_local = line.split(',')
            allowed_length = len(ExcelIO.export_columns)
            if len(data_local) < allowed_length:
                print(data_local)
                print(len(data_local), allowed_length)
                print("ERROR: Not enough values")
                continue
            print(data_local, allowed_length)
            while len(data_local) > allowed_length:
                data_local[allowed_length-1] = ','.join([data_local[allowed_length - 1], data_local[allowed_length]])
                print("APPENDING:", data_local[allowed_length])
                del data_local[allowed_length]
            print(data_local[allowed_length - 1])
            for index, dt in enumerate(data_local):
                try:
                    data_local[index] = int(dt)
                except ValueError:
                    continue
            self.data.append(tuple(data_local))
        self.file.close()

    def manipulate_file(self, data):
        self.file.close()
        self.file = open(self.file_path, 'w')
        self.file.writelines(data)
        self.file.close()

    def get_data(self):
        return self.data

    def get_text(self):
        try:
            return self.file.read().splitlines()
        except ValueError:
            self.file = open(self.file_path, 'r')
            return self.file.read().splitlines()


class ExcelIO:
    worksheet: Worksheet
    workbook: Workbook
    # export_columns = (1, 2, 3, 4, 6, 7, 9, 10, 12, 14, 15, 17, 18, 20, 22, 23)
    export_columns = tuple([n for n in range(1, 19)])

    def __init__(self):
        try:
            self.workbook = pyxl.load_workbook('data/output.xlsx')
        finally:
            pass
        self.name = str
        self.worksheet = ...

    def set_workbook(self, file_path):
        self.workbook = pyxl.load_workbook(file_path)
        self.name = file_path
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

    @staticmethod
    def get_cell(x: int, y: int, letter: str = '') -> str:
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                    'U', 'V', 'W', 'X', 'Y', 'Z']
        while x >= 1:
            if x <= 26:
                letter += alphabet[x - 1]
                x = 0
            else:
                letter += alphabet[int(x) // 26 - 1]
                x %= 26
        return letter + str(y)

    def modify_spreadsheet(self, data, start_checking=1):
        print(data)
        for i, inst in enumerate(data):
            export_row = start_checking
            while self.worksheet[self.get_cell(1, export_row)].value is not None:
                export_row += 1
            for n, column in enumerate(ExcelIO.export_columns):
                # print(self.get_cell(column, export_row))
                self.worksheet[self.get_cell(column, export_row)].value = inst[n]

        self.workbook.save(self.name)


class WidgetCustom:
    """Form widget base class to set default functions and values for all
    relevant widgets in the application."""

    default_font = ('Arial', 18)
    btn_pressed = False

    def __init__(self, widget, text=None):
        self.d_font = WidgetCustom.default_font
        self.displayable = True
        self.widget = widget
        self.text = text
        self.x_val = 0
        self.y_val = 0
        self.sup = self

    def place_custom(self, x_val: float = 0., y_val: float = 0.):

        """Method to make placing widgets on the screen easier.

        Parameters:
            x_val: rel-x position on window
            y_val: rel-y position on window"""

        # Updates x and y values
        self.x_val = x_val
        self.y_val = y_val
        try:
            self.widget.place(relx=x_val, rely=y_val, anchor=CENTER)
        except AttributeError:
            self.place(x_val, y_val, CENTER)
        finally:
            pass

    def get_value(self):

        """Method to return the widget's text value."""

        try:
            return self.widget.get()
        except AttributeError:
            return self.text.get()
        finally:
            pass

    def set_value(self, value):
        """Method to set a label's text value."""
        try:
            self.text.set(value)
        except AttributeError:
            pass

    def toggle_widget(self, desired_state=0):

        """Method to make toggling the widget's appearance easier."""

        if self.displayable and desired_state == 0:
            print(self.__str__(), 'Toggled off')
            self.widget.place_forget()
            # if type(self) == FormInput:
            #     self.border.place_forget()
        elif not self.displayable and desired_state == 1:
            print(self.__str__(), 'Toggled on')
            self.place_custom(self.x_val, self.y_val)
        else:
            print('Nothing happened to', self.__str__())
            return
        self.displayable = not self.displayable

    def place(self, *args):
        print(self, "'s ", *args, " unused.", sep='')
        return None


class LabelCustom(WidgetCustom):
    """Form label subclass of form widget.

    Creates a default Label widget based on settings for the form."""

    def __init__(self, window,
                 content='',
                 d_font_style=WidgetCustom.default_font,
                 d_width=25,
                 d_height=1,
                 d_border=2,
                 d_fg_color="white",
                 d_bg_color="black"):
        self.text = StringVar(master=window)
        self.widget = Label(window,
                            text=content,
                            font=d_font_style,
                            width=d_width,
                            height=d_height,
                            bd=d_border,
                            fg=d_fg_color,
                            bg=d_bg_color,
                            textvariable=self.text)
        self.text.set(content)
        super().__init__(self.widget, self.text)


class InputCustom(WidgetCustom):
    """Form input subclass of form widget.

    Creates a default Entry widget based on settings for the form."""

    def __init__(self, window,
                 d_font_style=WidgetCustom.default_font,
                 d_width=15,
                 d_fg_color="white",
                 d_bg_color="black"):
        self.border = Frame(window,
                            background="white",
                            width=d_width * App.dpi / 6.85,
                            height=2.5 * App.dpi / 6.85, )
        self.widget = Entry(window,
                            font=d_font_style,
                            width=d_width,
                            fg=d_fg_color,
                            bg=d_bg_color,
                            bd=0)

        super().__init__(self.widget)


class FileButtonCustom(WidgetCustom):
    """Form file button subclass of form widget.

    Creates a Button widget based on settings for the form. Configured to open
    a file dialog input system when clicked."""
    file_type: str
    widget: Button

    def __init__(self,
                 window, validated_cmd, invalidated_cmd=None, data_type='txt'):
        self.widget = Button(
            window,
            text="Browse File System",
            font=super().default_font,
            fg="black", bg="white",
            bd=0,
            command=self._browse_files)
        self.validated_cmd = validated_cmd
        self.invalidated_cmd = invalidated_cmd
        self.file_type = data_type
        self.widget.config(width=15)
        self.file_name = ''
        super().__init__(self.widget)

    def _browse_files(self):

        """Internal method to manage the filedialog window and resultantly
        attain the file path."""

        temp_name = filedialog.askopenfilename(title="Select a File",
                                               filetypes=(("File", str("." + self.file_type)),
                                                          ("all files", "*.*")))
        if self.file_name != '' and temp_name == '':
            return
        self.file_name = temp_name
        print(self.file_name)
        if self.file_name[-len(self.file_type):] == self.file_type:
            self.widget.configure(text='File inputted.')
            self.validated_cmd(self.file_name)
        else:
            self.invalidated_cmd()

    def get_file_lines(self):

        """Method to retrieve file lines separated as list items."""

        if self.file_name == '':
            return ['']
        file = open(self.file_name, "r", encoding="utf-8", errors="replace")
        lines = file.read().splitlines()
        file.close()
        return lines


class ScreenCustom:
    widgets: list[WidgetCustom]

    def __init__(self, screen_manager, *args):
        assert isinstance(screen_manager, ScreenManager)
        self.screen_manager = screen_manager
        self.widgets = []
        for arg in args:
            self.widgets.append(arg[0])
            self.widgets[-1].place_custom(arg[1], arg[2])
            self.widgets[-1].toggle_widget(0)

    def set_screen(self):
        if self.screen_manager.current_screen is not Ellipsis:
            print(self.screen_manager.current_screen)
            for widget in self.screen_manager.current_screen.widgets:
                widget.toggle_widget(0)
        for widget in self.widgets:
            widget.toggle_widget(1)
        self.screen_manager.current_screen = self


class ScreenManager:
    current_screen: ScreenCustom
    screens: dict[str, Any]

    def __init__(self):
        self.screens = {}
        self.current_screen = ...

    def get_screen(self, name):
        screen: ScreenCustom = self.screens[name]
        return screen

    def add_screen(self, name, *args: list[WidgetCustom, float, float]) -> None:
        self.screens[name] = ScreenCustom(self, *args)


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        import sys
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath("..")
    print(os.path.join(base_path, relative_path))
    return os.path.join(base_path, relative_path)


# Press the green button in the gutter to run the script.
# if __name__ == '__main__' and len(argv) >= 1:

def execute():
    am = App(*argv)
